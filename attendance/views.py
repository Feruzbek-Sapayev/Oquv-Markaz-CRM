from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from datetime import date
import calendar
from django.utils import timezone
from .models import AttendanceSession, Attendance
from .forms import AttendanceSessionForm, AttendanceFormSet
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from courses.models import Group, Enrollment
import datetime
from accounts.permissions import admin_required, teacher_required


UZ_MONTHS = [
    (1, 'yan'), (2, 'fev'), (3, 'mar'), (4, 'apr'), (5, 'may'), (6, 'iun'),
    (7, 'iul'), (8, 'avg'), (9, 'sen'), (10, 'okt'), (11, 'noy'), (12, 'dek')
]


@login_required
def session_list(request):
    group_id = request.GET.get('group', '')
    month_year = request.GET.get('month_year')
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    today = timezone.now().date()
    
    if month_year:
        try:
            month, year = map(int, month_year.split('-'))
        except (ValueError, TypeError):
            month, year = today.month, today.year
    else:
        try:
            month = int(month) if month else today.month
            year = int(year) if year else today.year
        except (ValueError, TypeError):
            month, year = today.month, today.year
        
    start_date = datetime.date(year, month, 1)
    # Get last day of month
    last_day = calendar.monthrange(year, month)[1]
    end_date = datetime.date(year, month, last_day)
    
    # Calculate prev/next month
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
        
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    groups = Group.objects.filter(is_active=True).select_related('course')
    
    # Filter groups based on role
    if request.user.is_teacher:
        groups = groups.filter(course__course_teachers__teacher__user=request.user).distinct()
    elif request.user.is_student:
        groups = groups.filter(enrollments__student__user=request.user, enrollments__is_active=True)
    elif not request.user.is_admin_role:
        groups = groups.none()
    selected_group = None
    matrix_data = []
    all_dates = []
    
    if group_id:
        selected_group = groups.filter(pk=group_id).first()
        # If student tries to access, they can only see their own attendance in the matrix
        # (The template needs adjustment to only show the student row if user.is_student)
        
    if selected_group:
        # Generate dates based on schedule
        potential_dates = [start_date + datetime.timedelta(days=i) for i in range(last_day)]
        
        # Filtering logic
        schedule = selected_group.days
        if schedule == Group.DayChoices.ODD:
            all_dates = [d for d in potential_dates if d.weekday() in [0, 2, 4]]
        elif schedule == Group.DayChoices.EVEN:
            all_dates = [d for d in potential_dates if d.weekday() in [1, 3, 5]]
        elif schedule == Group.DayChoices.WEEKEND:
            all_dates = [d for d in potential_dates if d.weekday() in [5, 6]]
        else: # daily or fallback
            all_dates = potential_dates
            
        # Get active students in this group
        enrollments = Enrollment.objects.filter(group=selected_group, is_active=True).select_related('student').order_by('student__last_name')
        
        # Get attendance records for this month
        attendance_records = Attendance.objects.filter(
            session__group=selected_group, session__date__range=[start_date, end_date]
        ).values('student_id', 'session__date', 'status')
        
        attendance_map = {}
        for r in attendance_records:
            sid, d, s = r['student_id'], r['session__date'], r['status']
            if sid not in attendance_map: attendance_map[sid] = {}
            attendance_map[sid][d] = s
            
        for e in enrollments:
            student_atts = []
            present_count = 0
            total_with_status = 0
            for d in all_dates:
                status = attendance_map.get(e.student.pk, {}).get(d)
                student_atts.append({'date': d, 'status': status})
                if status:
                    total_with_status += 1
                    if status == Attendance.Status.PRESENT:
                        present_count += 1
            
            percent = (present_count / total_with_status * 100) if total_with_status else 0
            
            if request.user.is_student:
                # If student, only add their own row to matrix_data
                if e.student.user == request.user:
                    matrix_data.append({
                        'student': e.student,
                        'attendances': student_atts,
                        'percentage': round(percent),
                    })
            else:
                matrix_data.append({
                    'student': e.student,
                    'attendances': student_atts,
                    'percentage': round(percent),
                })

    total_present = 0
    total_absent = 0
    total_records = 0
    
    for row in matrix_data:
        for att in row['attendances']:
            if att['status'] == Attendance.Status.PRESENT:
                total_present += 1
                total_records += 1
            elif att['status'] == Attendance.Status.ABSENT:
                total_absent += 1
                total_records += 1
                
    overall_percentage = round((total_present / total_records * 100)) if total_records > 0 else 0

    return render(request, 'attendance/session_list.html', {
        'groups': groups,
        'selected_group': group_id,
        'selected_group_obj': selected_group,
        'matrix_data': matrix_data,
        'all_dates': all_dates,
        'current_month': month,
        'current_year': year,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'uz_months': UZ_MONTHS,
        'full_uz_months': [
            (1, 'Yanvar'), (2, 'Fevral'), (3, 'Mart'), (4, 'Aprel'), (5, 'May'), (6, 'Iyun'),
            (7, 'Iyul'), (8, 'Avgust'), (9, 'Sentabr'), (10, 'Oktabr'), (11, 'Noyabr'), (12, 'Dekabr')
        ],
        'years': range(today.year - 2, today.year + 3),
        'today': today,
        'stats': {
            'present': total_present,
            'absent': total_absent,
            'percent': overall_percentage,
        }
    })


@teacher_required
def session_create(request, group_pk):
    group = get_object_or_404(Group, pk=group_pk)
    
    if not request.user.is_admin_role and not group.course.course_teachers.filter(teacher__user=request.user).exists():
        messages.error(request, "Faqat o'zingizning guruhlaringiz uchun davomat olishingiz mumkin!")
        return redirect('attendance:session_list')

    students = Enrollment.objects.filter(group=group, is_active=True).select_related('student')

    if request.method == 'POST':
        date = request.POST.get('date', str(timezone.now().date()))
        topic = request.POST.get('topic', '')

        session, created = AttendanceSession.objects.get_or_create(
            group=group, date=date,
            defaults={'topic': topic, 'created_by': request.user}
        )

        for enrollment in students:
            status = request.POST.get(f'status_{enrollment.student.pk}', 'absent')
            notes = request.POST.get(f'notes_{enrollment.student.pk}', '')
            Attendance.objects.update_or_create(
                session=session,
                student=enrollment.student,
                defaults={'status': status, 'notes': notes}
            )
        messages.success(request, f"{date} sanasi uchun davomat saqlandi!")
        return redirect('attendance:session_list')

    today = str(timezone.now().date())
    return render(request, 'attendance/session_create.html', {
        'group': group,
        'students': students,
        'today': today,
        'statuses': [s for s in Attendance.Status.choices if s[0] in ['present', 'absent']],
    })


@login_required
def session_detail(request, pk):
    session = get_object_or_404(
        AttendanceSession.objects.select_related('group__course', 'created_by'), pk=pk
    )
    records = session.records.select_related('student').order_by('student__last_name')
    present_count = records.filter(status=Attendance.Status.PRESENT).count()
    total_count = records.count()
    return render(request, 'attendance/session_detail.html', {
        'session': session,
        'records': records,
        'present_count': present_count,
        'total_count': total_count,
    })


@teacher_required
def session_delete(request, pk):
    session = get_object_or_404(AttendanceSession, pk=pk)
    group_pk = session.group.pk
    
    if not request.user.is_admin_role and not session.group.course.course_teachers.filter(teacher__user=request.user).exists():
        messages.error(request, "Sizda ushbu sessiyani o'chirish huquqi yo'q!")
        return redirect('attendance:session_list')
    if request.method == 'POST':
        session.delete()
        messages.success(request, 'Sessiya o\'chirildi!')
    return redirect('attendance:session_list')


@require_POST
@login_required
@teacher_required
def attendance_update_ajax(request):
    import json
    data = json.loads(request.body)
    student_id = data.get('student_id')
    date_str = data.get('date')
    status = data.get('status')
    group_id = data.get('group_id')
    
    date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    
    if date_obj > timezone.now().date():
        return JsonResponse({'status': 'error', 'message': 'Kelajakdagi sanaga davomat qilib bo\'lmaydi'}, status=400)

    session, created = AttendanceSession.objects.get_or_create(
        group_id=group_id, date=date_obj,
        defaults={'created_by': request.user}
    )
    
    # AJAX permission check
    if not request.user.is_admin_role and not session.group.course.course_teachers.filter(teacher__user=request.user).exists():
        return JsonResponse({'status': 'error', 'message': 'Faqat o\'z guruhingizga davomat qo\'yishingiz mumkin'}, status=403)
    
    if status in ['present', 'absent']:
        Attendance.objects.update_or_create(
            session=session, student_id=student_id,
            defaults={'status': status}
        )
    else:
        Attendance.objects.filter(session=session, student_id=student_id).delete()
        
    return JsonResponse({'status': 'success'})


@login_required
@teacher_required
def export_attendance_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    group_id = request.GET.get('group')
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    today = timezone.now().date()
    
    try:
        month = int(month) if month else today.month
        year = int(year) if year else today.year
    except ValueError:
        month, year = today.month, today.year
        
    if not group_id:
        return redirect('attendance:session_list')
        
    group = get_object_or_404(Group, pk=group_id)
    
    if not request.user.is_admin_role and not group.course.course_teachers.filter(teacher__user=request.user).exists():
        messages.error(request, "Faqat o'z guruhingiz davomatini yuklab olishingiz mumkin!")
        return redirect('attendance:session_list')
    start_date = datetime.date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = datetime.date(year, month, last_day)
    
    potential_dates = [start_date + datetime.timedelta(days=i) for i in range(last_day)]
    schedule = group.days
    if schedule == Group.DayChoices.ODD:
        all_dates = [d for d in potential_dates if d.weekday() in [0, 2, 4]]
    elif schedule == Group.DayChoices.EVEN:
        all_dates = [d for d in potential_dates if d.weekday() in [1, 3, 5]]
    elif schedule == Group.DayChoices.WEEKEND:
        all_dates = [d for d in potential_dates if d.weekday() in [5, 6]]
    else:
        all_dates = potential_dates

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Davomat"
    
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_dates) + 2)
    ws['A1'] = f"{group.name} Davomati ({start_date.strftime('%B %Y')})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Ism'] + [d.strftime('%d') for d in all_dates] + ['%']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    enrollments = Enrollment.objects.filter(group=group, is_active=True).select_related('student').order_by('student__last_name')
    
    attendance_records = Attendance.objects.filter(
        session__group=group, session__date__range=[start_date, end_date]
    ).values('student_id', 'session__date', 'status')
    
    attendance_map = {}
    for r in attendance_records:
        sid, d, s = r['student_id'], r['session__date'], r['status']
        if sid not in attendance_map: attendance_map[sid] = {}
        attendance_map[sid][d] = s
        
    row_idx = 3
    for e in enrollments:
        ws.cell(row=row_idx, column=1, value=str(e.student))
        present_count = 0
        total_with_status = 0
        for col_offset, d in enumerate(all_dates, 2):
            status = attendance_map.get(e.student.pk, {}).get(d)
            val = ''
            if status == 'present':
                val = 'Bor'
                present_count += 1
                total_with_status += 1
            elif status == 'absent':
                val = "Yo'q"
                total_with_status += 1
            
            cell = ws.cell(row=row_idx, column=col_offset, value=val)
            cell.alignment = Alignment(horizontal='center')
            if val == 'Bor': cell.font = Font(color='10B981', bold=True)
            elif val == "Yo'q": cell.font = Font(color='EF4444', bold=True)
            
        percent = (present_count / total_with_status * 100) if total_with_status else 0
        ws.cell(row=row_idx, column=len(all_dates) + 2, value=f"{round(percent)}%")
        ws.cell(row=row_idx, column=len(all_dates) + 2).alignment = Alignment(horizontal='center')
        row_idx += 1
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=davomat_{group.name}_{start_date.strftime("%Y_%m")}.xlsx'
    wb.save(response)
    return response
